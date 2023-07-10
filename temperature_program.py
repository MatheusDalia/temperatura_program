import tkinter as tk
from tkinter import messagebox
from tkinter import filedialog
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from tkinter.font import Font
from tkinter import Scrollbar
import json


class Application(tk.Tk):
    selected_csv = ""

    def __init__(self):
        super().__init__()
        self.title("Tabela de Max Temp e NHFT")
        self.geometry("900x400")

        self.term_carga = tk.BooleanVar()

        self.tipo_var = tk.StringVar()
        self.pavimentos_var = tk.IntVar()
        self.unidades_var = tk.IntVar()
        self.apps_var = tk.IntVar()
        self.quarto_var = tk.StringVar()

        self.current_pavimento = 1
        self.current_app = 1
        self.current_unidade = 1
        self.carga_resfr = 0

        self.pavimentos_data = []
        self.unidades_data = []
        self.apps_data = []

        self.unidades_apps = []  # List to hold app information for each unit

        self.create_widgets()

    def create_widgets(self):
        # CSV file input
        csv_label = tk.Label(self, text="VN File:")
        csv_label.pack()

        self.csv_entry = tk.Entry(self)
        self.csv_entry.pack()

        browse_button = tk.Button(self, text="Browse", command=self.browse_csv)
        browse_button.pack()

        # Tipo de UH label and radio buttons
        tipo_label = tk.Label(self, text="Tipo de UH?")
        tipo_label.pack()

        unifamiliar_radio = tk.Radiobutton(
            self, text="Unifamiliar", variable=self.tipo_var, value="Unifamiliar")
        unifamiliar_radio.pack()

        multifamiliar_radio = tk.Radiobutton(
            self, text="Multifamiliar", variable=self.tipo_var, value="Multifamiliar")
        multifamiliar_radio.pack()

        self.checkbox_var = tk.IntVar()

        # Checkbox for threshold selection
        threshold_frame = tk.Frame(self)
        threshold_frame.pack()

        self.threshold_var = tk.StringVar(
            value="28")  # Default threshold value

        intervalo1_checkbox = tk.Checkbutton(
            threshold_frame, text="Intervalo 1 - 18,0 °C < ToAPPa < 26,0 °C", variable=self.threshold_var, onvalue="26", offvalue="28")
        intervalo1_checkbox.pack(side="left")

        intervalo2_checkbox = tk.Checkbutton(
            threshold_frame, text="Intervalo 2 - ToAPP < 28,0 °C", variable=self.threshold_var, onvalue="28", offvalue="28")
        intervalo2_checkbox.pack(side="left")

        intervalo3_checkbox = tk.Checkbutton(
            threshold_frame, text="Intervalo 3 - ToAPP < 30,0 °C", variable=self.threshold_var, onvalue="30", offvalue="28")
        intervalo3_checkbox.pack(side="left")

        # Tipo de UH label and radio buttons
        carga_label = tk.Label(self, text="Contabiliza carga térmica?")
        carga_label.pack()

        carga_radio = tk.Radiobutton(
            self, text="Sim", variable=self.term_carga, value=True)
        carga_radio.pack()

        carga_radio = tk.Radiobutton(
            self, text="Não", variable=self.term_carga, value=False)
        carga_radio.pack()

        next_button = tk.Button(self, text="Next", command=self.on_next_button)
        next_button.pack()

        json_button = tk.Button(self, text="Upload JSON",
                                command=self.browse_json)
        json_button.pack()

    def browse_json(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json")])
        if file_path:
            try:
                with open(file_path) as json_file:
                    data = json.load(json_file)
                    if self.validate_json_data(data):
                        self.pavimentos_data = data
                        self.show_json_csv()
                    else:
                        messagebox.showerror("Error", "Invalid JSON file.")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def validate_json_data(self, data):
        required_keys = ["Nome do pavimento", "Quantas unidades", "unidades"]
        if isinstance(data, list):
            for pavimento in data:
                if isinstance(pavimento, dict) and all(key in pavimento for key in required_keys):
                    unidades = pavimento["unidades"]
                    if isinstance(unidades, list):
                        for unidade in unidades:
                            if not isinstance(unidade, dict):
                                return True
                            if "Nome da unidade" not in unidade or "Quantas APPs" not in unidade or "APPs" not in unidade:
                                return True
                            apps = unidade["APPs"]
                            if not isinstance(apps, list):
                                return True
                            for app in apps:
                                if not isinstance(app, dict) or "Codigo da APP" not in app or "Tipo de quarto" not in app or "Nome da APP" not in app:
                                    return True
                    else:
                        return True
                else:
                    return True
            return True
        return True

    def browse_csv(self):
        global selected_csv
        file_path = filedialog.askopenfilename(
            filetypes=[("CSV files", "*.csv")])
        self.csv_entry.delete(0, tk.END)  # Clear the current entry field
        # Set the selected file path in the entry field
        self.csv_entry.insert(0, file_path)
        selected_csv = self.csv_entry.get()

    def browse_carga(self):
        global selected_carga
        file_path = filedialog.askopenfilename(
            filetypes=[("CSV files", "*.csv")])
        self.carga_entry.delete(0, tk.END)  # Clear the current entry field
        # Set the selected file path in the entry field
        self.carga_entry.insert(0, file_path)
        selected_carga = self.carga_entry.get()

    def filter_data(self, csv_file, room_type):
        df = pd.read_csv(csv_file)  # Read the CSV file into a pandas DataFrame

        if room_type == "Quarto":
            # Remove lines where "SCH_OCUP_DORM:Schedule Value" column is not 1
            df = df[df['SCH_OCUP_DORM:Schedule Value [](Hourly) '] == 1]
        elif room_type == "Sala":
            # Remove lines where "SCH_OCUP_SALA:Schedule Value" column is not 0
            df = df[df['SCH_OCUP_SALA:Schedule Value [](Hourly)'] != 0]

        return df

    def get_max_temperature(self, df, key):
        # Find the column index for the specified key
        key_column_index = df.columns.get_loc(key)

        # Get the column for the specified key and extract the temperature values
        temperature_column = df.iloc[:, key_column_index].values[1:]

        # Find the maximum temperature
        max_temperature = max(temperature_column)

        # Round the max temperature to two decimal places
        max_temperature_rounded = round(max_temperature, 2)

        return max_temperature_rounded

    def get_min_temperature(self, df, key):
        # Find the column index for the specified key
        key_column_index = df.columns.get_loc(key)

        # Get the column for the specified key and extract the temperature values
        temperature_column = df.iloc[:, key_column_index].values[1:]

        # Find the minimum temperature
        min_temperature = min(temperature_column)

        # Round the min temperature to two decimal places
        min_temperature_rounded = round(min_temperature, 2)

        return min_temperature_rounded

    def get_nhft_value(self, df, key):
        # Find the column index for the specified key
        key_column_index = df.columns.get_loc(key)

        # Get the column for the specified key and extract the values
        value_column = df.iloc[:, key_column_index].values[0:]

        # Get the selected threshold value from the checkbox
        threshold = float(self.threshold_var.get())

        # Count the number of values less than the specified maximum value
        if (threshold == 26):

            subcount1 = sum(1 for value in value_column if value < 26)
            subcount2 = sum(1 for value in value_column if value < 18)
            count = subcount1 - subcount2

        else:
            count = sum(1 for value in value_column if value < 28)

        return count

    def joule_to_kwh(self, energy_in_joules):
        # Conversion factor: 1 J = 2.77778e-7 kWh
        joule_to_kwh_conversion_factor = 2.77778e-7
        energy_in_kwh = energy_in_joules * joule_to_kwh_conversion_factor
        return energy_in_kwh

    def carga_term(self, carga_filtered, filtered_data, codigo, codigo_solo):
        global carga_resfr
        # Rename the columns in filtered_data if there are duplicates
        filtered_data = filtered_data.add_suffix("_filtered")

        # Add the columns of filtered_data to carga_filtered
        carga_filtered = pd.concat([carga_filtered, filtered_data], axis=1)
        column_titles = carga_filtered.columns

        temp_threshold = float(self.threshold_var.get())
        temperature_column = carga_filtered[codigo_solo +
                                            ":Zone Operative Temperature [C](Hourly)_filtered"]
        # Filter rows where temperature is above the threshold
        if (temp_threshold == 26):
            filtered_rows1 = carga_filtered[temperature_column < 18]
            filtered_rows2 = carga_filtered[temperature_column > 26]
            filtered_rows = pd.concat([filtered_rows1, filtered_rows2])
        else:
            filtered_rows = carga_filtered[temperature_column > temp_threshold]

        if codigo in column_titles:
            # Calculate the sum of the values in the codigo_column
            codigo_column = filtered_rows[codigo]
            # Sum all the values in the filtered area
            total_sum = codigo_column.sum()

            if codigo_solo + " IDEAL LOADS AIR SYSTEM:Zone Ideal Loads Zone Total Heating Energy [J](Hourly)" in column_titles:
                # Get the additional column values
                additional_column = filtered_rows[codigo_solo +
                                                  " IDEAL LOADS AIR SYSTEM:Zone Ideal Loads Zone Total Heating Energy [J](Hourly)"]
                # Sum the values in the additional column
                carga_resfr = additional_column.sum()
                # Add the additional sum to the total sum
                total_sum += carga_resfr

            total_converted = total_sum / 3600000
            carga_resfr = 0
            return total_converted
        else:
            return False

    def on_next_button(self):
        csv_file = self.csv_entry.get()

        if not csv_file:
            messagebox.showerror("Error", "Please select a CSV file.")
            return

        try:
            # Read the CSV file into a pandas DataFrame
            df = pd.read_csv(csv_file)
        except pd.errors.EmptyDataError:
            messagebox.showerror(
                "Error", "The CSV file is empty or does not exist.")
            return
        except pd.errors.ParserError:
            messagebox.showerror("Error", "Failed to parse the CSV file.")
            return

        tipo = self.tipo_var.get()

        if self.term_carga.get() == True:
            self.destroy_widgets()
            self.carga_termica()
        else:

            if tipo == "Multifamiliar":
                self.show_pavimentos()
            elif tipo == "Unifamiliar":
                self.apps_var.set(1)  # Set the default value for units to 1
                self.current_pavimento = 1  # Reset the current pavimento
                self.on_pavimentos_next_button()

    def carga_termica(self):
        # Handle the redirection to the other page with a different CSV file input
        if self.term_carga.get() == True:
            self.destroy_widgets()

            carga_label = tk.Label(self, text="Carga Termica File:")
            carga_label.pack()

            self.carga_entry = tk.Entry(self)
            self.carga_entry.pack()

            browse_button = tk.Button(
                self, text="Browse", command=self.browse_carga)
            browse_button.pack()

            next_button = tk.Button(
                self, text="Next", command=self.show_pavimentos)
            next_button.pack()

        else:
            # Show error message if carga_entry is not selected
            if not self.carga_entry.get():
                messagebox.showerror(
                    "Error", "Please enter a Carga Termica value.")
                return

            self.destroy_widgets()

            pavimentos_label = tk.Label(self, text="Quantos pavimentos?")
            pavimentos_label.pack()

            pavimentos_entry = tk.Entry(self, textvariable=self.pavimentos_var)
            pavimentos_entry.pack()

            next_button = tk.Button(
                self, text="Next", command=self.on_pavimentos_next_button)
            next_button.pack()

    def show_pavimentos(self):
        if self.term_carga.get() == True:
            # Handle the redirection to the other page with a different CSV file input
            self.destroy_widgets()

            carga_label = tk.Label(self, text="Carga Termica File:")
            carga_label.pack()

            self.carga_entry = tk.Entry(self)
            self.carga_entry.pack()

            browse_button = tk.Button(
                self, text="Browse", command=self.browse_carga)
            browse_button.pack()

            tipo = self.tipo_var.get()

            if tipo == "Multifamiliar":
                self.show_pavimentos()
            elif tipo == "Unifamiliar":
                self.apps_var.set(1)  # Set the default value for units to 1
                self.current_pavimento = 1  # Reset the current pavimento
                self.on_pavimentos_next_button()

            next_button = tk.Button(
                self, text="Next", command=self.on_other_page_next_button)
            next_button.pack()
        else:

            self.destroy_widgets()

            pavimentos_label = tk.Label(self, text="Quantos pavimentos?")
            pavimentos_label.pack()

            pavimentos_entry = tk.Entry(self, textvariable=self.pavimentos_var)
            pavimentos_entry.pack()

            next_button = tk.Button(
                self, text="Next", command=self.on_pavimentos_next_button)
            next_button.pack()

    def on_other_page_next_button(self):
        carga_file = self.carga_entry.get()

        if not carga_file:
            messagebox.showerror("Error", "Please select a CSV file.")
            return

        try:
            # Read the CSV file into a pandas DataFrame
            df = pd.read_csv(carga_file)
        except pd.errors.EmptyDataError:
            messagebox.showerror(
                "Error", "The CSV file is empty or does not exist.")
            return
        except pd.errors.ParserError:
            messagebox.showerror("Error", "Failed to parse the CSV file.")
            return

        self.show_pavimentos()

    def on_pavimentos_next_button(self):
        pavimentos = self.pavimentos_var.get()
        self.pavimentos_var.set(0)

        if self.tipo_var.get() == "Unifamiliar":
            self.current_pavimento = 1  # Reset the current pavimento
            self.show_pavimento(self.current_pavimento, 1)  # Set total to 1
        else:
            self.show_pavimento(1, pavimentos)

    def show_pavimento(self, current, total):
        self.destroy_widgets()

        # Create a Canvas widget
        canvas = tk.Canvas(self, width=900, height=400)

        # Create a Scrollbar widget and associate it with the Canvas widget
        scrollbar = Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        # Pack the Scrollbar and Canvas widgets
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # Create a frame inside the Canvas widget to hold the widgets
        frame = tk.Frame(canvas)

        # Configure the Canvas to scroll the frame
        canvas.create_window((0, 0), window=frame, anchor="nw")

        # Configure the scrolling region
        frame.bind("<Configure>", lambda event: canvas.configure(
            scrollregion=canvas.bbox("all")))

        # Dictionary to store pavimento names and unidades entry variables
        pavimento_unidades_dict = {}
        for i in range(total):
            pavimento_label = tk.Label(
                frame, text=f"Nome do pavimento {i + 1}/{total}?")
            pavimento_label.grid(row=i, column=0, padx=10, pady=5)

            pavimento_entry = tk.Entry(frame)
            pavimento_entry.grid(row=i, column=1, padx=10, pady=5)

            unidades_label = tk.Label(frame, text="Quantas unidades?")
            unidades_label.grid(row=i, column=2, padx=10, pady=5)

            unidades_entry = tk.Entry(frame)
            unidades_entry.grid(row=i, column=3, padx=10, pady=5)

            # Add relation to the dictionary
            pavimento_unidades_dict[pavimento_entry] = unidades_entry

        next_button = tk.Button(frame, text="Next", command=lambda: self.on_pavimento_next_button(
            total, {pav.get(): uni.get() for pav, uni in pavimento_unidades_dict.items()}))
        next_button.grid(row=total, column=0, columnspan=4, padx=10, pady=5)

    def on_pavimento_next_button(self, total_pavimento, pavimento_unidades_dict):
        pavimentos = []

        for i in range(total_pavimento):
            pavimento = {}
            pavimento["Nome do pavimento"] = list(
                pavimento_unidades_dict.keys())[i]
            pavimento["Quantas unidades"] = list(
                pavimento_unidades_dict.values())[i]
            pavimento["unidades"] = []
            pavimentos.append(pavimento)
        self.show_unidade(
            pavimentos, total_pavimento)

    def show_unidade(self, pavimentos, total_pavimento):
        self.destroy_widgets()

        # Create a Canvas widget
        canvas = tk.Canvas(self, width=900, height=400)

        # Create a Scrollbar widget and associate it with the Canvas widget
        scrollbar = Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        # Pack the Scrollbar and Canvas widgets
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # Create a frame inside the Canvas widget to hold the widgets
        frame = tk.Frame(canvas)

        # Configure the Canvas to scroll the frame
        canvas.create_window((0, 0), window=frame, anchor="nw")

        # Configure the scrolling region
        frame.bind("<Configure>", lambda event: canvas.configure(
            scrollregion=canvas.bbox("all")))

        pavimentos_unidades = []  # List to hold unidades for each pavimento

        def on_next_button_click():
            for i, unidades in enumerate(pavimentos_unidades):
                pavimentos[i]["unidades"] = unidades
            self.show_app(total_pavimento, pavimentos)

        row_index = 0  # Track the current row index

        for i in range(total_pavimento):
            num_units = int(pavimentos[i]["Quantas unidades"])
            pavimento_name = pavimentos[i]["Nome do pavimento"]

            unidades = []  # Separate unidades list for each pavimento

            for unit in range(num_units):
                unidade_label = tk.Label(
                    frame, text=f"Pavimento {pavimento_name} | Unidade {unit + 1} | Qual o nome da unidade?")
                unidade_label.grid(row=row_index, column=0, sticky="w")

                unidade_entry = tk.Entry(frame)
                unidade_entry.grid(row=row_index, column=1)

                apps_label = tk.Label(frame, text="Quantas APP's?")
                apps_label.grid(row=row_index, column=2, sticky="w")

                apps_entry = tk.Entry(frame)
                apps_entry.grid(row=row_index, column=3)

                unidades.append({
                    "Nome da unidade": unidade_entry,
                    "Quantas APPs": apps_entry,
                    "APPs": []
                })

                row_index += 1  # Increment row index for the next row

            # Add unidades list to pavimentos_unidades
            pavimentos_unidades.append(unidades)

        def get_unidades_values():
            for unidades in pavimentos_unidades:
                for unidade in unidades:
                    unidade["Nome da unidade"] = unidade["Nome da unidade"].get()
                    unidade["Quantas APPs"] = unidade["Quantas APPs"].get()

        next_button = tk.Button(frame, text="Next", command=lambda: [
                                get_unidades_values(), on_next_button_click()])
        next_button.grid(row=row_index, column=0, columnspan=4)

    def show_app(self, total_pavimento, pavimentos):
        self.destroy_widgets()

        # Create a Canvas widget
        canvas = tk.Canvas(self, width=900, height=400)

        # Create a Scrollbar widget and associate it with the Canvas widget
        scrollbar = Scrollbar(self, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)

        # Pack the Scrollbar and Canvas widgets
        scrollbar.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)

        # Create a frame inside the Canvas widget to hold the widgets
        frame = tk.Frame(canvas)

        # Configure the Canvas to scroll the frame
        canvas.create_window((0, 0), window=frame, anchor="nw")

        # Configure the scrolling region
        frame.bind("<Configure>", lambda event: canvas.configure(
            scrollregion=canvas.bbox("all")))

        bold_font = Font(weight="bold")
        unidades_apps = []  # List to hold app information

        def on_next_button_click():
            for i in range(total_pavimento):
                num_units = int(pavimentos[i]["Quantas unidades"])

                for unit in range(num_units):
                    num_apps = int(
                        pavimentos[i]["unidades"][unit]['Quantas APPs'])
                    app_values = []  # List to hold app input values

                    for app in range(num_apps):
                        # Retrieve input values for each app
                        pavimento_code_value = pavimentos[i]["Nome do pavimento"]
                        unidade_code_value = pavimentos[i]["unidades"][unit]["Nome da unidade"]
                        app_code_value = unidades_apps[unit][app]["Codigo da APP"].get(
                        )
                        tipo_quarto_value = unidades_apps[unit][app]["Tipo de quarto"].get(
                        )
                        nome_app_value = unidades_apps[unit][app]["Nome da APP"].get(
                        )

                        app_values.append({
                            "Pavimento": pavimento_code_value,
                            "Unidade": unidade_code_value,
                            "Codigo da APP": app_code_value,
                            "Tipo de quarto": tipo_quarto_value,
                            "Nome da APP": nome_app_value
                        })

                    pavimentos[i]["unidades"][unit]["APPs"].append(app_values)

            self.pavimentos_data.append(pavimentos[0])
            self.show_output_button()

        row_index = 0  # Track the current row index

        for i in range(total_pavimento):
            num_units = int(pavimentos[i]["Quantas unidades"])
            pavimento_name = pavimentos[i]["Nome do pavimento"]

            unidades = pavimentos[i]["unidades"]
            unidades_apps = []  # List to hold app information for the current unit

            for unit in range(num_units):
                num_apps = int(pavimentos[i]["unidades"][unit]["Quantas APPs"])
                unidade_name = pavimentos[i]["unidades"][unit]["Nome da unidade"]
                app_data = []  # List to hold app information for the current unit

                for app in range(num_apps):
                    app_label = tk.Label(
                        frame, font=bold_font, text=f"Pavimento: {pavimento_name} | Unidade: {unidade_name} | APP: {app + 1} / {num_apps}")
                    app_label.grid(row=row_index, column=0,
                                   columnspan=4, sticky="w")

                    app_code_label = tk.Label(frame, text="Código da APP?")
                    app_code_label.grid(row=row_index + 1,
                                        column=0, sticky="w")

                    app_code_entry = tk.Entry(frame)
                    app_code_entry.grid(row=row_index + 1, column=1)

                    tipo_label = tk.Label(frame, text="Tipo de ambiente?")
                    tipo_label.grid(row=row_index + 1, column=2, sticky="w")

                    quarto_frame = tk.Frame(frame)
                    quarto_frame.grid(row=row_index + 1, column=3, sticky="w")

                    quarto_var = tk.StringVar()
                    quarto_checkbox = tk.Checkbutton(
                        quarto_frame, text="Quarto", variable=quarto_var, onvalue="Quarto")
                    quarto_checkbox.pack(side="left")

                    sala_checkbox = tk.Checkbutton(
                        quarto_frame, text="Sala", variable=quarto_var, onvalue="Sala")
                    sala_checkbox.pack(side="left")

                    nome_label = tk.Label(frame, text="Nome da APP?")
                    nome_label.grid(row=row_index + 1, column=4, sticky="w")

                    nome_entry = tk.Entry(frame)
                    nome_entry.grid(row=row_index + 1, column=5, columnspan=3)

                    empty_row_label = tk.Label(frame, text="", height=1)
                    empty_row_label.grid(row=row_index + 1, column=0, pady=4)

                    app_data.append({
                        "Pavimento": pavimento_name,
                        "Unidade": unidade_name,
                        "Codigo da APP": app_code_entry,
                        "Tipo de quarto": quarto_var,
                        "Nome da APP": nome_entry
                    })

                    row_index += 3

                # Append app_data to unidades_apps for the current unit
                unidades_apps.append(app_data)

            # Append unidades_apps to self.unidades_apps
            self.unidades_apps.append(unidades_apps)

        next_button = tk.Button(
            frame, text="Next", command=on_next_button_click)
        next_button.grid(row=row_index + 1, column=0, columnspan=4)

    def destroy_widgets(self):
        for widget in self.winfo_children():
            widget.destroy()

    def show_json_csv(self):
        self.destroy_widgets()
        csv_label = tk.Label(self, text="VN File:")
        csv_label.pack()

        self.csv_entry = tk.Entry(self)
        self.csv_entry.pack()

        browse_button = tk.Button(self, text="Browse", command=self.browse_csv)
        browse_button.pack()

        # Checkbox for threshold selection
        threshold_frame = tk.Frame(self)
        threshold_frame.pack()

        self.threshold_var = tk.StringVar(
            value="28")  # Default threshold value

        intervalo1_checkbox = tk.Checkbutton(
            threshold_frame, text="Intervalo 1 - 18,0 °C < ToAPPa < 26,0 °C", variable=self.threshold_var, onvalue="26", offvalue="28")
        intervalo1_checkbox.pack(side="left")

        intervalo2_checkbox = tk.Checkbutton(
            threshold_frame, text="Intervalo 2 - ToAPP < 28,0 °C", variable=self.threshold_var, onvalue="28", offvalue="28")
        intervalo2_checkbox.pack(side="left")

        intervalo3_checkbox = tk.Checkbutton(
            threshold_frame, text="Intervalo 3 - ToAPP < 30,0 °C", variable=self.threshold_var, onvalue="30", offvalue="28")
        intervalo3_checkbox.pack(side="left")

        # Tipo de UH label and radio buttons
        carga_label = tk.Label(self, text="Contabiliza carga térmica?")
        carga_label.pack()

        carga_radio = tk.Radiobutton(
            self, text="Sim", variable=self.term_carga, value=True)
        carga_radio.pack()

        carga_radio = tk.Radiobutton(
            self, text="Não", variable=self.term_carga, value=False)
        carga_radio.pack()

        csv_next_button = tk.Button(
            self, text="Next", command=self.redirect_next_button)
        csv_next_button.pack()

    def redirect_next_button(self):
        if self.term_carga.get():
            self.pre_output_carga_json()
        else:
            self.show_output_button()

    def pre_output_carga_json(self):
        self.destroy_widgets()

        carga_label = tk.Label(self, text="Carga Termica File:")
        carga_label.pack()

        self.carga_entry = tk.Entry(self)
        self.carga_entry.pack()

        browse_button = tk.Button(
            self, text="Browse", command=self.browse_carga)
        browse_button.pack()

        csv_next_button = tk.Button(
            self, text="Next", command=self.show_output_button)
        csv_next_button.pack()

    def show_output_button(self):
        self.destroy_widgets()

        output_button = tk.Button(
            self, text="Output", command=self.generate_output)
        output_button.pack()

        download_button = tk.Button(
            self, text="Download JSON", command=self.download_json)
        download_button.pack()

    def download_json(self):
        file_path = filedialog.asksaveasfilename(
            defaultextension=".json", filetypes=[("JSON files", "*.json")])
        if not file_path:
            return

        with open(file_path, "w") as json_file:
            json.dump(self.pavimentos_data, json_file)

    def generate_output(self):
        if not self.pavimentos_data:
            messagebox.showinfo("Information", "No data to export!")
            return

        data = []

        filepath = selected_csv
        if (self.term_carga.get() == True):
            cargapath = selected_carga

        for pavimento in self.pavimentos_data:
            for unidade in pavimento["unidades"]:
                for app in unidade["APPs"][0]:
                    filtered_data = self.filter_data(
                        filepath, app["Tipo de quarto"])
                    if (self.term_carga.get() == True):
                        carga_filtered_data = self.filter_data(
                            cargapath, app["Tipo de quarto"])
                    max_temp = self.get_max_temperature(
                        filtered_data, app["Codigo da APP"] + ':Zone Operative Temperature [C](Hourly)')
                    value_count = self.get_nhft_value(
                        filtered_data, app["Codigo da APP"] + ':Zone Operative Temperature [C](Hourly)')
                    min_temp = self.get_min_temperature(
                        filtered_data, app["Codigo da APP"] + ':Zone Operative Temperature [C](Hourly)')
                    if (self.term_carga.get() == True):
                        carga = self.carga_term(carga_filtered_data, filtered_data,
                                                app["Codigo da APP"] + ' IDEAL LOADS AIR SYSTEM:Zone Ideal Loads Zone Total Cooling Energy [J](Hourly)', app["Codigo da APP"])
                    if app["Tipo de quarto"] == "Quarto":
                        phft_value = (value_count / 3650) * 100
                    else:
                        phft_value = (value_count / 2920) * 100

                    if (self.term_carga.get() == True):

                        row = {
                            "Pavimento": app["Pavimento"],
                            "Unidade": app["Unidade"],
                            "Código": app["Codigo da APP"],
                            "Nome": app["Nome da APP"],
                            "Tipo de ambiente": app["Tipo de quarto"],
                            "MIN TEMP": min_temp,
                            "MAX TEMP": max_temp,
                            "NHFT": value_count,
                            "PHFT": phft_value,
                            "CARGA RESF": carga - (carga_resfr/3600000),
                            "CARGA AQUE": carga_resfr/3600000,
                            "CARGA TERM": carga
                        }
                    else:
                        row = {
                            "Pavimento": app["Pavimento"],
                            "Unidade": app["Unidade"],
                            "Código": app["Codigo da APP"],
                            "Nome": app["Nome da APP"],
                            "Tipo de ambiente": app["Tipo de quarto"],
                            "MIN TEMP": min_temp,
                            "MAX TEMP": max_temp,
                            "NHFT": value_count,
                            "PHFT": phft_value,
                        }

                    data.append(row)

        output_data = pd.DataFrame(data)

        messagebox.showinfo(
            "Information", "Click OK to export the data to Excel.")

        # Ask the user to choose the directory and filename
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])

        if file_path:
            export_to_excel(output_data, file_path)


def export_to_excel(data, filename):
    if data is not None:
        data.to_excel(filename, index=False)

        # Open the Excel file
        wb = openpyxl.load_workbook(filename)
        sheet = wb.active

        # Apply styles to column headers
        pavimento_header = sheet.cell(row=1, column=1)
        unidade_header = sheet.cell(row=1, column=2)
        codigo_temp_header = sheet.cell(row=1, column=3)
        nome_header = sheet.cell(row=1, column=4)
        ambiente_header = sheet.cell(row=1, column=5)
        max_temp_header = sheet.cell(row=1, column=6)
        nhft_header = sheet.cell(row=1, column=7)
        phft_header = sheet.cell(row=1, column=8)
        carga_header = sheet.cell(row=1, column=9)

        pavimento_header.fill = PatternFill(
            start_color="f8cbad", end_color="f8cbad", fill_type="solid")  # Pink background color
        unidade_header.fill = PatternFill(
            start_color="f8cbad", end_color="f8cbad", fill_type="solid")  # Pink background color
        codigo_temp_header.fill = PatternFill(
            start_color="f8cbad", end_color="f8cbad", fill_type="solid")  # Pink background color
        nome_header.fill = PatternFill(
            start_color="f8cbad", end_color="f8cbad", fill_type="solid")  # Pink background color
        ambiente_header.fill = PatternFill(
            start_color="f8cbad", end_color="f8cbad", fill_type="solid")  # Pink background color

        max_temp_header.fill = PatternFill(
            start_color="8faadc", end_color="8faadc", fill_type="solid")  # Blue background color
        nhft_header.fill = PatternFill(
            start_color="a9d18e", end_color="a9d18e", fill_type="solid")  # Green background color
        # Adjust cell width to fit content
        phft_header.fill = PatternFill(
            start_color="8faadc", end_color="8faadc", fill_type="solid")  # Blue background color
        carga_header.fill = PatternFill(
            start_color="a9d18e", end_color="a9d18e", fill_type="solid")  # Green background color
        # Adjust cell width to fit content
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except TypeError:
                    pass
            adjusted_width = (max_length + 2) * 1.2  # Add some extra width
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Save the modified Excel file
        wb.save(filename)

        messagebox.showinfo("Information", f"Data exported to {filename}!")
    else:
        messagebox.showinfo("Information", "No data to export!")


if __name__ == "__main__":
    app = Application()
    app.mainloop()
