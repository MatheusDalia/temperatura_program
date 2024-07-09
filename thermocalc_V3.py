import tkinter as tk
from tkinter import messagebox, filedialog
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from tkinter.font import Font
from tkinter import Scrollbar
import json


class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Tabela de Max Temp e NHFT")
        self.geometry("900x400")
        self.selected_csv = ""
        self.setup_vars()
        self.create_widgets()

    def setup_vars(self):
        self.term_carga = tk.BooleanVar()
        self.tipo_var = tk.StringVar()
        self.pavimentos_var = tk.IntVar()
        self.unidades_var = tk.IntVar()
        self.apps_var = tk.IntVar()
        self.quarto_var = tk.StringVar()
        self.current_pavimento = 1
        self.carga_resfr = 0
        self.pavimentos_data = []
        self.unidades_apps = []

    def create_widgets(self):
        self.create_input_widgets()
        self.create_radio_buttons()
        self.create_checkboxes()
        self.create_json_widgets()
        tk.Button(self, text="Next", command=self.on_next_button).pack()

    def create_input_widgets(self):
        tk.Label(self, text="VN File:").pack()
        self.csv_entry = tk.Entry(self)
        self.csv_entry.pack()
        tk.Button(self, text="Browse", command=self.browse_csv).pack()

    def create_radio_buttons(self):
        tk.Label(self, text="Tipo de UH?").pack()
        tk.Radiobutton(self, text="Unifamiliar",
                       variable=self.tipo_var, value="Unifamiliar").pack()
        tk.Radiobutton(self, text="Multifamiliar",
                       variable=self.tipo_var, value="Multifamiliar").pack()

    def create_checkboxes(self):
        threshold_frame = tk.Frame(self)
        threshold_frame.pack()
        self.threshold_var = tk.StringVar(value="28")
        intervals = [
            ("Intervalo 1 - 18,0 °C < ToAPPa < 26,0 °C", "26"),
            ("Intervalo 2 - ToAPP < 28,0 °C", "28"),
            ("Intervalo 3 - ToAPP < 30,0 °C", "30"),
        ]
        for text, value in intervals:
            tk.Checkbutton(
                threshold_frame,
                text=text,
                variable=self.threshold_var,
                onvalue=value,
                offvalue="28",
            ).pack(side="left")

    def create_json_widgets(self):
        tk.Label(self, text="Contabiliza carga térmica?").pack()
        tk.Radiobutton(self, text="Sim",
                       variable=self.term_carga, value=True).pack()
        tk.Radiobutton(self, text="Não",
                       variable=self.term_carga, value=False).pack()
        tk.Button(self, text="Upload JSON", command=self.browse_json).pack()
        # Add entry widget for Excel file
        self.excel_entry = tk.Entry(self)
        self.excel_entry.pack()
        # Button to upload Excel file
        excel_button = tk.Button(
            self, text="Upload Excel", command=self.browse_excel)
        excel_button.pack()

    def browse_csv(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("CSV files", "*.csv")])
        self.csv_entry.delete(0, tk.END)
        self.csv_entry.insert(0, file_path)
        self.selected_csv = self.csv_entry.get()

    def browse_excel(self):
        excel_file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx")])
        if excel_file_path:
            self.excel_entry.delete(0, tk.END)  # Clear any previous entry
            self.excel_entry.insert(0, excel_file_path)

    def browse_json(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json")])
        if file_path:
            try:
                with open(file_path) as json_file:
                    data = json.load(json_file)
                    if self.validate_json_data(data):
                        self.pavimentos_data = data
                        self.show_json_csv()
                    else:
                        messagebox.showerror("Error", "Invalid JSON file.")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def validate_json_data(self, data):
        required_keys = ["Nome do pavimento", "Quantas unidades", "unidades"]
        return all(self.is_valid_pavimento(pavimento, required_keys) for pavimento in data) if isinstance(data, list) else False

    def is_valid_pavimento(self, pavimento, required_keys):
        if not (isinstance(pavimento, dict) and all(key in pavimento for key in required_keys)):
            return False
        unidades = pavimento["unidades"]
        return all(self.is_valid_unidade(unidade) for unidade in unidades) if isinstance(unidades, list) else False

    def is_valid_unidade(self, unidade):
        required_keys = ["Nome da unidade", "Quantas APPs", "APPs"]
        if not (isinstance(unidade, dict) and all(key in unidade for key in required_keys)):
            return False
        apps = unidade["APPs"]
        return all(self.is_valid_app(app) for app in apps) if isinstance(apps, list) else False

    def is_valid_app(self, app):
        required_keys = ["Codigo da APP", "Tipo de quarto", "Nome da APP"]
        return isinstance(app, dict) and all(key in app for key in required_keys)

    def filter_data(self, csv_file, room_type):
        df = pd.read_csv(csv_file)
        df.columns = df.columns.str.strip()
        print("Available columns:", df.columns)
        filters = {
            "Quarto": "SCH_OCUP_DORM:Schedule Value [](Hourly)",
            "Sala": "SCH_OCUP_SALA:Schedule Value [](Hourly)",
        }
        column_name = filters.get(room_type, "")
        if column_name not in df.columns:
            messagebox.showerror(
                "Error", f"Column '{column_name}' not found in the CSV file.")
            return pd.DataFrame()  # Return an empty DataFrame to handle the error gracefully
        return df[df[column_name] == 1] if room_type == "Quarto" else df[df[column_name] != 0]

    def get_max_temperature(self, df, key):
        key = key.strip()
        return round(df[key].max(), 2)

    def get_min_temperature(self, df, key):
        key = key.strip()
        return round(df[key].min(), 2)

    def get_nhft_value(self, df, key):
        key = key.strip()
        value_column = df[key]
        threshold = float(self.threshold_var.get())
        if threshold == 26:
            count = ((value_column < 26) & (value_column >= 18)).sum()
        else:
            count = (value_column < threshold).sum()
        return count

    def joule_to_kwh(self, energy_in_joules):
        return energy_in_joules * 2.77778e-7

    def carga_term(self, carga_filtered, filtered_data, codigo, codigo_solo):
        global carga_resfr
        filtered_data = filtered_data.add_suffix("_filtered")
        carga_filtered = pd.concat([carga_filtered, filtered_data], axis=1)
        temperature_column = carga_filtered[
            f"{codigo_solo}:Zone Operative Temperature [C](Hourly)_filtered"]
        temp_threshold = float(self.threshold_var.get())

        filtered_rows = (
            pd.concat([carga_filtered[temperature_column < 18],
                      carga_filtered[temperature_column > 26]])
            if temp_threshold == 26
            else carga_filtered[temperature_column > temp_threshold]
        )

        if codigo in carga_filtered.columns:
            carga_column = filtered_rows[codigo]
            total_sum = carga_column.sum()
            heating_col = f"{codigo_solo} IDEAL LOADS AIR SYSTEM:Zone Ideal Loads Zone Total Heating Energy [J](Hourly)"
            if heating_col in carga_filtered.columns:
                heating_sum = filtered_rows[heating_col].sum()
                total_sum += heating_sum
                carga_resfr = heating_sum
            total_kwh = total_sum / 3600000
            carga_resfr = 0
            return total_kwh

        return False

    def on_next_button(self):
        csv_file = self.selected_csv
        if not csv_file:
            messagebox.showerror("Error", "Please select a CSV file.")
            return
        try:
            pd.read_csv(csv_file)
        except (pd.errors.EmptyDataError, pd.errors.ParserError):
            messagebox.showerror("Error", "Failed to parse the CSV file.")
            return
        room_type = self.quarto_var.get()
        df = self.filter_data(csv_file, room_type)
        if df.empty:
            return  # Exit if there was an error filtering the data
        column_name = "SCH_OCUP_DORM:Schedule Value [](Hourly)"
        max_temp = self.get_max_temperature(df, column_name)
        nhft_value = self.get_nhft_value(df, column_name)
        messagebox.showinfo(
            "Results", f"Max Temp: {max_temp}\nNHFT Value: {nhft_value}")
        pavimento_name = self.pavimentos_data[self.current_pavimento -
                                              1]["Nome do pavimento"]
        unidade_name = self.unidades_apps[self.current_unidade -
                                          1]["Nome da unidade"]
        app_name = self.unidades_apps[self.current_unidade -
                                      1]["APPs"][self.current_app - 1]["Nome da APP"]
        carga_resfr = self.carga_term(df, df, column_name, column_name)
        pavimento_data = {
            "Pavimento": pavimento_name,
            "Unidade": unidade_name,
            "APP": app_name,
            "Max Temp": max_temp,
            "NHFT": nhft_value,
            "Carga de Resfriamento": carga_resfr,
        }
        self.save_to_excel(pavimento_data)

    def show_json_csv(self):
        if not self.pavimentos_data:
            messagebox.showerror("Error", "No data available.")
            return
        pavimento = self.pavimentos_data[0]
        pavimento_name = pavimento["Nome do pavimento"]
        self.unidades_apps = pavimento["unidades"]
        self.current_pavimento = 1
        self.show_unidade_options()

    def show_unidade_options(self):
        unidades = self.unidades_apps
        unidade_options = [unidade["Nome da unidade"] for unidade in unidades]
        self.unidade_var = tk.StringVar()
        tk.OptionMenu(self, self.unidade_var, *unidade_options).pack()

    def save_to_excel(self, data):
        excel_path = self.excel_entry.get().strip()
        if not excel_path:
            messagebox.showerror(
                "Error", "Please provide a valid Excel file path.")
            return
        try:
            wb = openpyxl.load_workbook(excel_path)
            sheet = wb.active
            row = sheet.max_row + 1
            for col, key in enumerate(data, start=1):
                sheet.cell(row=row, column=col, value=data[key])
            wb.save(excel_path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save to Excel: {str(e)}")


if __name__ == "__main__":
    app = Application()
    app.mainloop()
