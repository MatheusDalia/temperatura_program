import tkinter as tk
from tkinter import messagebox, filedialog
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from tkinter.font import Font
from tkinter import Scrollbar
import json


class Application(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Tabela de Max Temp e NHFT")
        self.geometry("900x400")
        self.selected_csv = ""
        self.setup_vars()
        self.create_widgets()

    def setup_vars(self):
        self.term_carga = tk.BooleanVar()
        self.tipo_var = tk.StringVar()
        self.pavimentos_var = tk.IntVar()
        self.unidades_var = tk.IntVar()
        self.apps_var = tk.IntVar()
        self.quarto_var = tk.StringVar()
        self.current_pavimento = 1
        self.current_app = 1
        self.current_unidade = 1
        self.carga_resfr = 0
        self.pavimentos_data = []
        self.unidades_data = []
        self.apps_data = []
        self.unidades_apps = []

    def create_widgets(self):
        tk.Label(self, text="VN File:").pack()
        self.csv_entry = tk.Entry(self)
        self.csv_entry.pack()
        tk.Button(self, text="Browse", command=self.browse_csv).pack()
        self.create_radio_buttons()
        self.create_checkboxes()
        self.create_carga_widgets()
        tk.Button(self, text="Upload JSON", command=self.browse_json).pack()
        tk.Button(self, text="Next", command=self.on_next_button).pack()

    def create_radio_buttons(self):
        tk.Label(self, text="Tipo de UH?").pack()
        tk.Radiobutton(self, text="Unifamiliar",
                       variable=self.tipo_var, value="Unifamiliar").pack()
        tk.Radiobutton(self, text="Multifamiliar",
                       variable=self.tipo_var, value="Multifamiliar").pack()

    def create_checkboxes(self):
        threshold_frame = tk.Frame(self)
        threshold_frame.pack()
        self.threshold_var = tk.StringVar(value="28")
        intervals = [("Intervalo 1 - 18,0 °C < ToAPPa < 26,0 °C", "26"),
                     ("Intervalo 2 - ToAPP < 28,0 °C", "28"),
                     ("Intervalo 3 - ToAPP < 30,0 °C", "30")]
        for text, value in intervals:
            tk.Checkbutton(threshold_frame, text=text, variable=self.threshold_var,
                           onvalue=value, offvalue="28").pack(side="left")

    def create_carga_widgets(self):
        tk.Label(self, text="Contabiliza carga térmica?").pack()
        tk.Radiobutton(self, text="Sim",
                       variable=self.term_carga, value=True).pack()
        tk.Radiobutton(self, text="Não",
                       variable=self.term_carga, value=False).pack()
        tk.Label(self, text="Excel File:").pack()
        self.excel_entry = tk.Entry(self)
        self.excel_entry.pack()
        tk.Button(self, text="Upload Excel", command=self.browse_excel).pack()

    def browse_csv(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("CSV files", "*.csv")])
        self.csv_entry.delete(0, tk.END)
        self.csv_entry.insert(0, file_path)
        self.selected_csv = self.csv_entry.get()

    def browse_excel(self):
        excel_file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx")])
        if excel_file_path:
            self.excel_entry.delete(0, tk.END)
            self.excel_entry.insert(0, excel_file_path)

    def browse_json(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("JSON files", "*.json")])
        if file_path:
            try:
                with open(file_path) as json_file:
                    data = json.load(json_file)
                    if self.validate_json_data(data):
                        self.pavimentos_data = data
                        self.show_json_csv()
                    else:
                        messagebox.showerror("Error", "Invalid JSON file.")
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def validate_json_data(self, data):
        required_keys = ["Nome do pavimento", "Quantas unidades", "unidades"]
        if not isinstance(data, list):
            return False
        for pavimento in data:
            if not all(key in pavimento for key in required_keys):
                return False
            if not isinstance(pavimento["unidades"], list):
                return False
            for unidade in pavimento["unidades"]:
                if not all(key in unidade for key in ["Nome da unidade", "Quantas APPs", "APPs"]):
                    return False
                if not isinstance(unidade["APPs"], list):
                    return False
                for app in unidade["APPs"]:
                    if not all(key in app for key in ["Codigo da APP", "Tipo de quarto", "Nome da APP"]):
                        return False
        return True

    def filter_data(self, csv_file, room_type):
        df = pd.read_csv(csv_file)
        df.columns = df.columns.str.strip()
        print(df.columns)  # Para verificar os nomes das colunas

        column_mapping = {
            "Quarto": "SCH_OCUP_DORM:Schedule Value [](Hourly)",
            "Sala": "SCH_OCUP_SALA:Schedule Value [](Hourly)",
            "Misto": "SCH_OCUP_MISTO:Schedule Value [](Hourly)"
        }

        # Verifica se a coluna necessária existe no DataFrame
        column_name = column_mapping.get(room_type)
        if column_name and column_name not in df.columns:
            messagebox.showerror(
                "Error", f"Column '{column_name}' not found in CSV.")
            return pd.DataFrame()  # Retorna um DataFrame vazio em caso de erro

        filters = {
            "Quarto": df[column_mapping["Quarto"]] == 1 if column_mapping["Quarto"] in df.columns else pd.Series([False] * len(df)),
            "Sala": df[column_mapping["Sala"]] != 0 if column_mapping["Sala"] in df.columns else pd.Series([False] * len(df)),
            "Misto": df[column_mapping["Misto"]] != 0 if column_mapping["Misto"] in df.columns else pd.Series([False] * len(df))
        }

        return df[filters.get(room_type, df)]

    def get_max_temperature(self, df, key):
        key = key.strip()
        max_temperature = df[key].max()
        return round(max_temperature, 2)

    def get_min_temperature(self, df, key):
        key = key.strip()
        min_temperature = df[key].min()
        return round(min_temperature, 2)

    def get_nhft_value(self, df, key):
        key = key.strip()
        value_column = df[key]
        threshold = float(self.threshold_var.get())
        if threshold == 26:
            count = ((value_column < 26) & (value_column >= 18)).sum()
        else:
            count = (value_column < threshold).sum()
        return count

    def on_next_button(self):
        csv_file = self.selected_csv
        excel_file = self.excel_entry.get()
        if not csv_file:
            messagebox.showerror("Error", "Please select a CSV file.")
            return
        if excel_file:
            if self.term_carga.get():
                self.carga_termica_excel(excel_file)
            else:
                self.process_excel_data(excel_file)
        else:
            try:
                pd.read_csv(csv_file)
            except (pd.errors.EmptyDataError, pd.errors.ParserError):
                messagebox.showerror("Error", "Failed to parse the CSV file.")

    def process_excel_data(self, excel_file):
        workbook = openpyxl.load_workbook(excel_file)
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            pavimento, unidade, codigo, nome, tipo_ambiente = row[:5]
            filtered_data = self.filter_data(self.selected_csv, tipo_ambiente)
            max_temp = self.get_max_temperature(
                filtered_data, f"{codigo.strip()}:Zone Operative Temperature [C](Hourly)")
            nhft = self.get_nhft_value(
                filtered_data, f"{codigo.strip()}:Zone Operative Temperature [C](Hourly)")
            min_temp = self.get_min_temperature(
                filtered_data, f"{codigo.strip()}:Zone Operative Temperature [C](Hourly)")
            phft = self.calculate_phft(tipo_ambiente, nhft)
            row_data = {
                "Pavimento": pavimento,
                "Unidade": unidade,
                "Código": codigo,
                "Nome": nome,
                "Tipo de ambiente": tipo_ambiente,
                "MIN TEMP": min_temp,
                "MAX TEMP": max_temp,
                "NHFT": nhft,
                "PHFT": phft
            }
            data.append(row_data)
        output_data = pd.DataFrame(data)
        self.export_to_excel(output_data, filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")]))

    def calculate_phft(self, tipo_ambiente, nhft):
        days_per_year = {"Quarto": 3650, "Misto": 6570, "Sala": 2920}
        return (nhft / days_per_year.get(tipo_ambiente, 365)) * 100

    def export_to_excel(self, data, filename):
        if data is not None:
            data.to_excel(filename, index=False)
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
            styles = {
                "Pavimento": "f8cbad",
                "Unidade": "f8cbad",
                "Código": "f8cbad",
                "Nome": "f8cbad",
                "Tipo de ambiente": "f8cbad",
                "MAX TEMP": "8faadc",
                "NHFT": "a9d18e",
                "PHFT": "8faadc",
                "CARGA TERM": "a9d18e"
            }
            for col, fill_color in styles.items():
                cell = sheet[f"{col}1"]
                cell.fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid")
                sheet.column_dimensions[cell.column_letter].width = max(
                    len(cell.value), 15)
            workbook.save(filename)
            messagebox.showinfo("Information", f"Data exported to {filename}!")


if __name__ == "__main__":
    app = Application()
    app.mainloop()
